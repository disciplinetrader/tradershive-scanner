"""Deterministic tests for scan-level Situational Awareness."""

import pytest

from app.engine.situation import SituationEngine
from app.features.situation import SituationFeature
from app.models.decision import DecisionAction
from app.models.facts import Facts
from app.models.market import MarketRegime
from app.models.sector import SectorProfile, SectorRotation
from app.models.situation import (
    Aggression,
    MoneyFlow,
    PositionSizingGuidance,
    RiskEnvironment,
    TradingBias,
)
from tests.test_decision_engine import _evaluate, _market, _sector
from tests.test_scanner_report import build_test_scanner


def _sectors() -> tuple[SectorProfile, ...]:
    """Build ten ranked sectors spanning all rotation states."""
    profiles = []
    rotations = (
        SectorRotation.LEADING,
        SectorRotation.LEADING,
        SectorRotation.IMPROVING,
        SectorRotation.IMPROVING,
        SectorRotation.LEADING,
        SectorRotation.WEAKENING,
        SectorRotation.WEAKENING,
        SectorRotation.LAGGING,
        SectorRotation.LAGGING,
        SectorRotation.LAGGING,
    )
    for index, rotation in enumerate(rotations, 1):
        profile = _sector(score=100 - index * 5, rotation=rotation)
        sector_facts = profile.facts.model_copy(update={"name": f"Sector {index}"})
        profiles.append(profile.model_copy(update={"rank": index, "facts": sector_facts}))
    return tuple(profiles)


def _analyze(facts: Facts, regime: MarketRegime, score: float = 75):
    """Analyze one repeated candidate universe under a requested regime."""
    market = _market(regime, score)
    decision = _evaluate(facts, market, _sectors()[0])
    return SituationEngine().analyze(
        market,
        _sectors(),
        (facts.rs_profile,) * 10,
        (facts.stock_profile,) * 10,
        (facts.setup_profile,) * 10,
        (facts.risk_profile,) * 10,
        (decision,) * 10,
    )


@pytest.mark.parametrize(
    ("regime", "bias", "aggression"),
    [
        (MarketRegime.HEALTHY_BULL, TradingBias.LONG_ONLY, Aggression.VERY_HIGH),
        (MarketRegime.BULL, TradingBias.LONG_BIAS, Aggression.HIGH),
        (MarketRegime.RANGE, TradingBias.NEUTRAL, Aggression.LOW),
        (MarketRegime.BEAR, TradingBias.CASH, Aggression.VERY_LOW),
        (MarketRegime.CAPITULATION, TradingBias.CASH, Aggression.VERY_LOW),
        (MarketRegime.RECOVERY, TradingBias.LONG_BIAS, Aggression.MEDIUM),
    ],
)
def test_regime_postures(
    bullish_facts: Facts,
    regime: MarketRegime,
    bias: TradingBias,
    aggression: Aggression,
) -> None:
    """Every required regime should map to an explicit operating posture."""
    profile = _analyze(bullish_facts, regime, 95 if "Bull" in regime.value else 50)
    assert profile.market_regime == regime
    assert profile.trading_bias == bias
    assert profile.aggression == aggression
    assert profile.recommended_setup_types
    if bias == TradingBias.CASH:
        assert profile.position_sizing_guidance == PositionSizingGuidance.MINIMAL
        assert profile.recommended_maximum_open_positions == 0
        assert profile.maximum_risk_per_trade == 0


def test_healthy_bull_is_risk_on_with_full_guidance(bullish_facts: Facts) -> None:
    """Strong breadth and leadership should permit the highest operating posture."""
    profile = _analyze(bullish_facts, MarketRegime.HEALTHY_BULL, 95)
    assert profile.money_flow == MoneyFlow.RISK_ON
    assert profile.risk_environment == RiskEnvironment.LOW
    assert profile.position_sizing_guidance == PositionSizingGuidance.FULL
    assert profile.recommended_maximum_open_positions == 12
    assert profile.maximum_risk_per_trade == 1
    assert "Broad market confirmation" in profile.reasons


def test_weak_breadth_reduces_aggression_and_turns_risk_off(bullish_facts: Facts) -> None:
    """Weak participation should override an otherwise healthy regime posture."""
    market = _market(MarketRegime.HEALTHY_BULL, 75)
    breadth = market.breadth.model_copy(
        update={"percentage_above_ema50": 30, "percentage_above_ema20": 35}
    )
    market = market.model_copy(update={"breadth": breadth})
    decision = _evaluate(bullish_facts, market)
    profile = SituationEngine().analyze(
        market,
        _sectors(),
        (bullish_facts.rs_profile,),
        (bullish_facts.stock_profile,),
        (bullish_facts.setup_profile,),
        (bullish_facts.risk_profile,),
        (decision,),
    )
    assert profile.aggression == Aggression.HIGH
    assert profile.money_flow == MoneyFlow.RISK_OFF
    assert "Weak breadth requires reduced aggression" in profile.warnings


def test_sector_rotation_and_high_volatility_are_exposed(bullish_facts: Facts) -> None:
    """Leadership lists and extreme VIX should directly affect awareness output."""
    market = _market(MarketRegime.HEALTHY_BULL, 80)
    volatility = market.volatility.model_copy(update={"india_vix": 35})
    market = market.model_copy(update={"volatility": volatility})
    decision = _evaluate(bullish_facts, market)
    sectors = _sectors()
    profile = SituationEngine().analyze(
        market,
        sectors,
        (bullish_facts.rs_profile,),
        (bullish_facts.stock_profile,),
        (bullish_facts.setup_profile,),
        (bullish_facts.risk_profile,),
        (decision,),
    )
    assert profile.risk_environment == RiskEnvironment.EXTREME
    assert profile.aggression == Aggression.HIGH
    assert profile.sector_leadership.top_sectors == tuple(f"Sector {i}" for i in range(1, 6))
    assert "Sector 3" in profile.sector_leadership.improving_sectors
    assert "Sector 6" in profile.sector_leadership.weakening_sectors
    assert SituationFeature().summarize(profile)[0] == "Environment: Healthy Bull"


def test_scanner_and_excel_expose_shared_situation(rising_frame, tmp_path) -> None:
    """Every result should share one profile and reports should put it first."""
    from openpyxl import load_workbook

    from app.reports.excel import generate_excel_report

    results = build_test_scanner(rising_frame).scan(["STRONG", "WEAK"])
    assert all(result.situation_profile is not None for result in results)
    assert results[0].situation_profile == results[1].situation_profile
    report = generate_excel_report(results, tmp_path / "situation.xlsx")
    workbook = load_workbook(report)
    assert workbook.sheetnames[0] == "Situation Summary"
    assert workbook["Situation Summary"]["A2"].value == "Market Regime"
    assert workbook["Situation Summary"]["A3"].value == "Breadth State"
    assert "CPR Environment" in {
        workbook["Situation Summary"].cell(row=row, column=1).value
        for row in range(1, workbook["Situation Summary"].max_row + 1)
    }
    assert results[0].decision_profile.action in DecisionAction
