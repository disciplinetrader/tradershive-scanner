"""Integration-style tests for orchestration and Excel reporting."""

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from app.core.constants import FEATURE_WEIGHTS
from app.data.loader import DataLoader
from app.engine.registry import FeatureRegistry
from app.engine.scorer import Scorer
from app.features.avwap import AVWAPFeature
from app.features.breadth import BreadthFeature
from app.features.cpr import CPRFeature
from app.features.market import MarketFeature
from app.features.momentum import MomentumFeature
from app.features.relative_strength import RelativeStrengthFeature
from app.features.risk import RiskFeature
from app.features.sector import SectorFeature
from app.features.setup import SetupFeature
from app.features.stock import StockFeature
from app.features.trend import TrendFeature
from app.features.volatility import VolatilityFeature
from app.features.volume import VolumeFeature
from app.providers.base import MarketDataProvider
from app.reports.excel import TOP_HEADERS, generate_excel_report
from app.scanner.scanner import Scanner


class SymbolProvider(MarketDataProvider):
    """Return deterministic frames with symbol-specific strength."""

    def __init__(self, frame: pd.DataFrame) -> None:
        """Store baseline history."""
        self._frame = frame

    def history(self, symbol: str, period: str) -> pd.DataFrame:
        """Return weaker history for the designated WEAK ticker."""
        frame = self._frame.copy()
        if symbol == "WEAK.NS":
            frame.loc[:, "Close"] = frame["Close"].iloc[::-1].to_numpy()
            frame.loc[:, "Open"] = frame["Close"]
            frame.loc[:, "High"] = frame["Close"] * 1.01
            frame.loc[:, "Low"] = frame["Close"] * 0.99
        return frame


def build_test_scanner(frame: pd.DataFrame) -> Scanner:
    """Construct the complete scanner with an in-memory provider."""
    registry = FeatureRegistry(
        [
            MarketFeature(),
            BreadthFeature(),
            CPRFeature(),
            AVWAPFeature(),
            SectorFeature(),
            StockFeature(),
            SetupFeature(),
            RiskFeature(),
            TrendFeature(),
            RelativeStrengthFeature(),
            MomentumFeature(),
            VolumeFeature(),
            VolatilityFeature(),
        ]
    )
    return Scanner(DataLoader(SymbolProvider(frame)), Scorer(registry, FEATURE_WEIGHTS), "^NSEI")


def test_scanner_ranks_results_and_report_is_styled_without_changing_data(
    rising_frame: pd.DataFrame, tmp_path: Path
) -> None:
    """The report should be usable and preserve the scanner's exact payload."""
    results = build_test_scanner(rising_frame).scan(["WEAK", "STRONG"])
    assert [result.rank for result in results] == [1, 2]
    assert results[0].symbol == "STRONG.NS"
    assert results[0].final_score > results[1].final_score

    report = generate_excel_report(results, tmp_path / "scanner.xlsx")
    workbook = load_workbook(report)
    assert workbook.sheetnames == [
        "Situation Summary",
        "Top 20",
        "Detailed Candidates",
        "Advanced Setup Details",
        "Volume Intelligence",
        "Sector and Industry Leadership",
        "CPR",
        "AVWAP",
        "Risk",
        "Validation Summary",
    ]

    top = workbook["Top 20"]
    headers = [top.cell(10, column).value for column in range(1, len(TOP_HEADERS) + 1)]
    assert headers == TOP_HEADERS
    assert top["A1"].value == "TRADERSHIVE"
    assert top["A2"].value.startswith("EOD MOMENTUM SCANNER")
    assert top.freeze_panes == "C11"
    assert top.auto_filter.ref == "A10:W12"
    assert list(top.tables) == ["Top20Results"]
    assert len(top.conditional_formatting) >= 5
    assert top["B11"].value == results[0].symbol
    assert top["D11"].value == results[0].decision_score
    assert top["M11"].value == results[0].facts.relative_strength_percentile
    assert top["Q11"].value == "N/A"
    assert top["D11"].number_format == "0.00"
    assert top["F11"].number_format == "0%"
    assert top["L11"].number_format == "0.00x"
    assert top["M11"].number_format == '0.0"%"'

    details = workbook["Detailed Candidates"]
    detail_headers = [cell.value for cell in details[5]]
    assert "Decision Reasons" in detail_headers
    assert "RS Profile" in detail_headers
    assert "AVWAP Anchors" in detail_headers
    assert "Target 5R" in detail_headers
    assert details.freeze_panes == "C6"
    assert details.tables["DetailedCandidates"].ref.endswith("7")

    for sheet in workbook.worksheets:
        assert sheet.sheet_view.showGridLines is False
        assert sheet.page_setup.orientation == "landscape"

    assert results[0].facts.relative_strength_percentile == 99
    assert results[1].facts.relative_strength_percentile == 0


def test_excel_report_rejects_empty_results(tmp_path: Path) -> None:
    """An empty workbook must not masquerade as a valid scan report."""
    try:
        generate_excel_report([], tmp_path / "empty.xlsx")
    except ValueError as error:
        assert str(error) == "Cannot generate a report without results"
    else:
        raise AssertionError("Expected empty report generation to fail")
