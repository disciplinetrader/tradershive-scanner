"""Deterministic historical validation engine tests."""

from datetime import date

import pandas as pd
import pytest
from openpyxl import load_workbook

from app.core.v11_config import ScannerProfileConfig, ScannerProfileName
from app.validation.engine import (
    ForwardOutcomeEvaluator,
    HistoricalReplayService,
    ValidationReportBuilder,
)
from app.validation.models import (
    AmbiguityPolicy,
    EntryModel,
    HistoricalCandidate,
    HistoricalTrade,
    TradeOutcome,
    ValidationConfig,
)
from app.validation.reports import write_excel_report, write_json_report, write_trade_ledger


def candidate(**updates: object) -> HistoricalCandidate:
    """Create a complete historical candidate with overridable fields."""
    values = {
        "scan_date": date(2025, 1, 1),
        "symbol": "TEST.NS",
        "rank": 1,
        "decision_score": 96,
        "grade": "A+",
        "action": "BUY",
        "setup_type": "VCP",
        "market_regime": "Healthy Bull",
        "sector": "Defence",
        "rs_percentile": 97,
        "volume_state": "Professional Accumulation",
        "cpr_state": "Trending",
        "avwap_state": "Strong Support",
        "risk_grade": "A",
        "pivot_price": 101,
        "risk_entry_price": 101,
        "stop_price": 99,
    }
    values.update(updates)
    return HistoricalCandidate.model_validate(values)


def bars(highs: list[float], lows: list[float], closes: list[float] | None = None) -> pd.DataFrame:
    """Create post-scan EOD bars."""
    close = closes or [(high + low) / 2 for high, low in zip(highs, lows, strict=True)]
    return pd.DataFrame(
        {"Open": [101.0] * len(highs), "High": highs, "Low": lows, "Close": close, "Volume": 1_000},
        index=pd.bdate_range("2025-01-02", periods=len(highs)),
    )


def test_pivot_not_triggered() -> None:
    """A pivot entry cannot exist when no future bar traded through it."""
    result = ForwardOutcomeEvaluator().evaluate(
        candidate(pivot_price=110),
        bars([105, 106], [100, 101]),
        ValidationConfig(entry_model=EntryModel.PIVOT_TRIGGER),
    )
    assert result.outcome == TradeOutcome.NOT_TRIGGERED
    assert result.entry_price is None


@pytest.mark.parametrize("target", [2, 3, 4, 5])
def test_targets_two_through_five(target: int) -> None:
    """Each configured R target exits at its exact objective price."""
    result = ForwardOutcomeEvaluator().evaluate(
        candidate(),
        bars([101, 101 + target * 2], [100, 100]),
        ValidationConfig(entry_model=EntryModel.RISK_PROFILE, target_r=target),
    )
    assert result.outcome == TradeOutcome.WIN
    assert result.realized_r == target
    assert getattr(result, f"target_{target}r_hit")


def test_stop_and_same_bar_ambiguity_are_conservative() -> None:
    """The default policy stops out a candle that also touches the target."""
    result = ForwardOutcomeEvaluator().evaluate(
        candidate(), bars([107], [98]), ValidationConfig(entry_model=EntryModel.RISK_PROFILE)
    )
    assert result.stop_hit and result.ambiguity_flag
    assert result.realized_r == -1


def test_maximum_holding_period_exits_at_close() -> None:
    """Unresolved trades leave at the final configured session close."""
    result = ForwardOutcomeEvaluator().evaluate(
        candidate(),
        bars([102] * 10, [100] * 10, [101.5] * 10),
        ValidationConfig(entry_model=EntryModel.RISK_PROFILE, maximum_holding_sessions=5),
    )
    assert result.holding_sessions == 5
    assert result.exit_price == 101.5
    assert result.realized_r == 0.25


def test_breakdowns_top_n_and_walk_forward() -> None:
    """Score, setup, regime, Top-N, and walk-forward cohorts stay separated."""
    evaluator = ForwardOutcomeEvaluator()
    config = ValidationConfig(entry_model=EntryModel.RISK_PROFILE)
    trades = [
        evaluator.evaluate(candidate(rank=rank, decision_score=score), bars([107], [100]), config)
        for rank, score in ((1, 96), (6, 92), (15, 82), (20, 65))
    ]
    builder = ValidationReportBuilder()
    report = builder.build(trades, config)
    assert "95-100" in report.breakdowns["score_buckets"].groups
    assert "VCP" in report.breakdowns["setup_types"].groups
    assert "Healthy Bull" in report.breakdowns["market_regimes"].groups
    assert report.top_n_comparison["Top 5"].number_of_candidates == 1
    assert report.top_n_comparison["Top 20"].number_of_candidates == 4
    walk = builder.walk_forward(
        trades, [(date(2024, 1, 1), date(2024, 12, 31), date(2025, 1, 1), date(2025, 12, 31))]
    )
    assert walk.periods[0].in_sample_metrics.number_of_candidates == 0
    assert walk.periods[0].out_of_sample_metrics.number_of_candidates == 4


def test_empty_results_and_missing_data_warnings() -> None:
    """Empty validation sets remain serializable and warnings reduce data quality."""
    report = ValidationReportBuilder().build(
        [], ValidationConfig(), ("Missing benchmark data", "Delisted-stock omissions possible")
    )
    assert report.metrics.number_of_candidates == 0
    assert report.data_quality.score < 100
    assert len(report.data_quality.warnings) == 2


def test_no_lookahead_and_duplicate_scan_prevention() -> None:
    """Replay passes only bars dated on or before each scan date."""
    observed: list[date] = []

    class FakeScanner:
        def __init__(self, histories: dict[str, pd.DataFrame]) -> None:
            self.histories = histories

        def scan(self, symbols: object, sectors: object) -> list[object]:
            observed.extend(
                pd.Timestamp(frame.index[-1]).date() for frame in self.histories.values()
            )
            return []

    frame = bars([101] * 300, [99] * 300)
    service = HistoricalReplayService(FakeScanner)  # type: ignore[arg-type]
    scan_date = pd.Timestamp(frame.index[270]).date()
    service.replay({"TEST.NS": frame}, [scan_date], ValidationConfig())
    assert observed == [scan_date]
    with pytest.raises(ValueError, match="Duplicate"):
        service.replay({"TEST.NS": frame}, [scan_date, scan_date], ValidationConfig())


def test_report_outputs_have_required_sheets(tmp_path) -> None:
    """JSON, CSV, and the ten-sheet workbook preserve the ledger."""
    config = ValidationConfig(entry_model=EntryModel.RISK_PROFILE)
    trade = ForwardOutcomeEvaluator().evaluate(candidate(), bars([107], [100]), config)
    report = ValidationReportBuilder().build([trade], config)
    write_json_report(report, tmp_path / "report.json")
    write_trade_ledger(report.trades, tmp_path / "ledger.csv")
    path = write_excel_report(report, tmp_path / "report.xlsx")
    assert (tmp_path / "report.json").exists() and (tmp_path / "ledger.csv").exists()
    assert load_workbook(path).sheetnames == [
        "Summary",
        "Score Buckets",
        "Setup Types",
        "Market Regimes",
        "Sectors",
        "Annual Results",
        "Top-N Comparison",
        "Advanced Setups",
        "Volume Signatures",
        "Market Pressure",
        "Industry Groups",
        "Scanner Profiles",
        "Feature Analysis",
        "Trade Ledger",
        "Walk Forward",
        "Data Quality Warnings",
    ]


def test_target_first_policy_is_explicit() -> None:
    """Sensitivity analysis may select target-first while retaining the flag."""
    result = ForwardOutcomeEvaluator().evaluate(
        candidate(),
        bars([107], [98]),
        ValidationConfig(
            entry_model=EntryModel.RISK_PROFILE, ambiguity_policy=AmbiguityPolicy.TARGET_FIRST
        ),
    )
    assert result.outcome == TradeOutcome.WIN and result.ambiguity_flag


def test_next_close_excludes_entry_candle_and_requires_later_session() -> None:
    """Entry-session extremes cannot affect a trade entered at that session's close."""
    frame = bars([120, 102], [90, 100], [101, 101.5])
    result = ForwardOutcomeEvaluator().evaluate(
        candidate(), frame, ValidationConfig(entry_model=EntryModel.NEXT_CLOSE)
    )
    incomplete = ForwardOutcomeEvaluator().evaluate(
        candidate(), frame.iloc[:1], ValidationConfig(entry_model=EntryModel.NEXT_CLOSE)
    )
    assert not result.stop_hit and not result.target_2r_hit
    assert result.maximum_favorable_excursion == 1
    assert result.maximum_adverse_excursion == 1
    assert incomplete.outcome == TradeOutcome.INCOMPLETE


@pytest.mark.parametrize(
    ("policy", "outcome", "stop_hit", "target_hit"),
    [
        (AmbiguityPolicy.STOP_FIRST, TradeOutcome.LOSS, True, False),
        (AmbiguityPolicy.TARGET_FIRST, TradeOutcome.WIN, False, True),
        (AmbiguityPolicy.FLAG_ONLY, TradeOutcome.INVALID, False, False),
    ],
)
def test_same_bar_policy_controls_target_credit(
    policy: AmbiguityPolicy,
    outcome: TradeOutcome,
    stop_hit: bool,
    target_hit: bool,
) -> None:
    """Ambiguous candles credit targets only under an explicit target-first policy."""
    result = ForwardOutcomeEvaluator().evaluate(
        candidate(),
        bars([112], [98]),
        ValidationConfig(
            entry_model=EntryModel.RISK_PROFILE,
            target_r=5,
            ambiguity_policy=policy,
        ),
    )
    assert result.outcome == outcome and result.ambiguity_flag
    assert result.stop_hit is stop_hit
    assert all(getattr(result, f"target_{multiple}r_hit") is target_hit for multiple in range(2, 6))


def test_feature_trigger_rate_uses_group_denominator() -> None:
    """Unequal feature cohorts calculate trigger rate within each cohort."""
    evaluator = ForwardOutcomeEvaluator()
    config = ValidationConfig(entry_model=EntryModel.RISK_PROFILE)
    triggered = evaluator.evaluate(candidate(setup_type="VCP"), bars([107], [100]), config)
    untriggered = HistoricalTrade(
        candidate=candidate(setup_type="VCP"), outcome=TradeOutcome.NOT_TRIGGERED
    )
    other = evaluator.evaluate(candidate(setup_type="Flat Base"), bars([107], [100]), config)
    report = ValidationReportBuilder().build([triggered, untriggered, other], config)
    vcp = report.feature_analysis["setup_types:VCP"]
    flat = report.feature_analysis["setup_types:Flat Base"]
    assert vcp.trigger_rate == 50
    assert flat.trigger_rate == 100
    assert vcp.prevalence == pytest.approx(66.67)


def test_replay_passes_profile_and_point_in_time_industry_mapping() -> None:
    """Each replay uses the selected policy and latest mapping known on its scan date."""
    observed: list[tuple[str, str]] = []

    class CapturingScanner:
        def __init__(self, histories: dict[str, pd.DataFrame]) -> None:
            self.histories = histories

        def scan(
            self,
            symbols: object,
            sectors: object,
            industries: dict[str, str],
            profile: ScannerProfileConfig,
        ) -> list[object]:
            observed.append((profile.name.value, industries["TEST.NS"]))
            return []

    frame = bars([101] * 300, [99] * 300)
    scan_date = pd.Timestamp(frame.index[270]).date()
    old_date = pd.Timestamp(frame.index[265]).date()
    mappings = {
        old_date: {"TEST.NS": "Legacy Group"},
        scan_date: {"TEST.NS": "Current Group"},
    }
    service = HistoricalReplayService(CapturingScanner)  # type: ignore[arg-type]
    service.replay(
        {"TEST.NS": frame},
        [scan_date],
        ValidationConfig(),
        industries_by_date=mappings,
        scanner_profile=ScannerProfileConfig(name=ScannerProfileName.IPO_LEADER),
    )
    service.replay(
        {"TEST.NS": frame},
        [old_date],
        ValidationConfig(),
        industries_by_date=mappings,
        scanner_profile=ScannerProfileConfig(name=ScannerProfileName.FIRST_PULLBACK),
    )
    assert observed == [("IPO Leader", "Current Group"), ("First Pullback", "Legacy Group")]


def test_replay_warns_when_point_in_time_industry_mapping_is_missing() -> None:
    """A future-only mapping cannot silently leak into an earlier replay."""

    class Scanner:
        def __init__(self, histories: dict[str, pd.DataFrame]) -> None:
            self.histories = histories

        def scan(self, *args: object) -> list[object]:
            return []

    frame = bars([101] * 300, [99] * 300)
    scan_date = pd.Timestamp(frame.index[270]).date()
    future_date = pd.Timestamp(frame.index[280]).date()
    _, warnings = HistoricalReplayService(Scanner).replay(  # type: ignore[arg-type]
        {"TEST.NS": frame},
        [scan_date],
        ValidationConfig(),
        industries_by_date={future_date: {"TEST.NS": "Future Group"}},
    )
    assert any("point-in-time industry mapping unavailable" in warning for warning in warnings)
