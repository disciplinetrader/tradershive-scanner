"""Professional, presentation-only Excel output for scanner results."""

from collections.abc import Iterable, Sequence
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, DataBarRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from app.models.stock_result import StockResult

CHARCOAL = "1F2328"
BLACK = "111315"
YELLOW = "F4C430"
WHITE = "FFFFFF"
LIGHT_GREY = "F3F4F6"
MID_GREY = "D9DDE3"
DARK_GREY = "5B616A"
GREEN = "C6EFCE"
DARK_GREEN = "548235"
PALE_YELLOW = "FFF2CC"
ORANGE = "FCE4D6"
RED = "FFC7CE"
FONT_NAME = "Aptos"
THIN_BORDER = Border(
    left=Side(style="thin", color=MID_GREY),
    right=Side(style="thin", color=MID_GREY),
    top=Side(style="thin", color=MID_GREY),
    bottom=Side(style="thin", color=MID_GREY),
)

TOP_HEADERS = [
    "Rank",
    "Symbol",
    "Action",
    "Decision Score",
    "Grade",
    "Confidence",
    "Best Setup",
    "Setup Grade",
    "Pivot",
    "Entry",
    "Stop",
    "Available R",
    "RS Percentile",
    "Sector",
    "Sector Rank",
    "Industry Group",
    "Industry Rank",
    "Volume Grade",
    "Volume State",
    "CPR State",
    "AVWAP State",
    "Risk Grade",
    "Warnings",
]


def _text(value: Any, default: str = "N/A") -> str:
    """Return enum/string display text without leaking ``None`` into Excel."""
    if value is None or value == "":
        return default
    return str(getattr(value, "value", value))


def _joined(values: Iterable[Any]) -> str:
    """Join explanatory values into a readable cell."""
    return " | ".join(str(getattr(value, "value", value)) for value in values)


def _configure_sheet(sheet: Worksheet, title_rows: str = "1:4") -> None:
    """Apply workbook-wide viewing and printing defaults."""
    sheet.sheet_view.showGridLines = False
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.autoPageBreaks = False
    sheet.oddHeader.center.text = f"&B{sheet.title}"
    sheet.oddFooter.left.text = "TradersHIVE EOD Momentum Scanner"
    sheet.oddFooter.center.text = "Page &P of &N"
    sheet.oddFooter.right.text = "Generated &D &T"
    sheet.print_title_rows = title_rows


def _title(sheet: Worksheet, title: str, subtitle: str, columns: int) -> None:
    """Write a compact branded title area above a sheet's data."""
    end = get_column_letter(max(columns, 2))
    sheet.merge_cells(f"A1:{end}1")
    sheet["A1"] = title
    sheet["A1"].font = Font(name=FONT_NAME, size=20, bold=True, color=YELLOW)
    sheet["A1"].fill = PatternFill("solid", fgColor=BLACK)
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 30
    sheet.merge_cells(f"A2:{end}2")
    sheet["A2"] = subtitle
    sheet["A2"].font = Font(name=FONT_NAME, size=10, color=WHITE)
    sheet["A2"].fill = PatternFill("solid", fgColor=CHARCOAL)
    sheet["A2"].alignment = Alignment(vertical="center")


def _header(cell: Any) -> None:
    """Style one table or section header cell."""
    cell.font = Font(name=FONT_NAME, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=CHARCOAL)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = THIN_BORDER


def _table(
    sheet: Worksheet,
    start_row: int,
    headers: Sequence[str],
    rows: Sequence[Sequence[Any]],
    name: str,
    *,
    freeze: bool = True,
) -> int:
    """Write a filtered, striped table and return its last row."""
    for column, header in enumerate(headers, 1):
        sheet.cell(start_row, column, header)
        _header(sheet.cell(start_row, column))
    for row_index, row in enumerate(rows, start_row + 1):
        for column, value in enumerate(row, 1):
            cell = sheet.cell(row_index, column, value)
            cell.font = Font(name=FONT_NAME, size=10, color=CHARCOAL)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top")
            if (row_index - start_row) % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=LIGHT_GREY)
    last_row = start_row + max(len(rows), 1)
    if not rows:
        for column in range(1, len(headers) + 1):
            sheet.cell(last_row, column, "N/A" if column == 1 else "")
    reference = f"A{start_row}:{get_column_letter(len(headers))}{last_row}"
    table = Table(displayName=name, ref=reference)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)
    sheet.auto_filter.ref = reference
    if freeze:
        sheet.freeze_panes = f"A{start_row + 1}"
    return last_row


def _widths(sheet: Worksheet, headers: Sequence[str], long_columns: set[str] | None = None) -> None:
    """Set bounded widths suitable for laptop viewing."""
    long_columns = long_columns or set()
    for index, header in enumerate(headers, 1):
        if header in long_columns:
            width = 48
        elif header in {"Symbol", "Sector", "Industry Group", "Best Setup", "Volume State"}:
            width = max(14, min(24, len(header) + 5))
        else:
            width = max(11, min(19, len(header) + 2))
        sheet.column_dimensions[get_column_letter(index)].width = width


def _format_columns(
    sheet: Worksheet, start_row: int, last_row: int, headers: Sequence[str]
) -> None:
    """Apply semantic number formats and alignment by column name."""
    price_names = {
        "Pivot",
        "Entry",
        "Stop",
        "Invalidation",
        "Nearest Resistance",
        "Target 2R",
        "Target 3R",
        "Target 4R",
        "Target 5R",
    }
    confidence_names = {
        "Confidence",
        "Decision Confidence",
        "Setup Confidence",
        "Volume Confidence",
        "Risk Confidence",
        "Event Confidence",
    }
    r_names = {"Available R", "Stop ATR"}
    percent_100_names = {name for name in headers if "%" in name or "Percentile" in name}
    score_names = {name for name in headers if "Score" in name}
    left_names = {
        "Symbol",
        "Sector",
        "Industry Group",
        "Best Setup",
        "Setup Type",
        "Volume State",
        "Volume Signature",
        "CPR State",
        "AVWAP State",
        "Reasons",
        "Warnings",
        "RS Profile",
        "AVWAP Anchors",
    }
    for column, name in enumerate(headers, 1):
        for row in range(start_row + 1, last_row + 1):
            cell = sheet.cell(row, column)
            if name in price_names:
                cell.number_format = "#,##0.00"
            elif name in confidence_names:
                cell.number_format = "0%"
            elif name in r_names:
                cell.number_format = "0.00x"
            elif name in percent_100_names:
                cell.number_format = '0.0"%"'
            elif name in score_names:
                cell.number_format = "0.00"
            if name in left_names:
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=name in {"Reasons", "Warnings", "RS Profile", "AVWAP Anchors"},
                )
            else:
                cell.alignment = Alignment(horizontal="center", vertical="top")


def _kpi(sheet: Worksheet, row: int, column: int, label: str, value: Any) -> None:
    """Render one two-column KPI card."""
    label_cell = sheet.cell(row, column, label)
    value_cell = sheet.cell(row + 1, column, value)
    for cell in (label_cell, value_cell):
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    label_cell.fill = PatternFill("solid", fgColor=CHARCOAL)
    label_cell.font = Font(name=FONT_NAME, size=9, bold=True, color=WHITE)
    value_cell.fill = PatternFill("solid", fgColor=LIGHT_GREY)
    value_cell.font = Font(name=FONT_NAME, size=11, bold=True, color=BLACK)


def _top_row(result: StockResult) -> list[Any]:
    """Create the concise dashboard row without changing report values."""
    decision = result.decision_profile
    facts = result.facts
    return [
        result.rank,
        result.symbol,
        _text(decision.action if decision else None, ""),
        result.decision_score,
        _text(decision.grade if decision else None, ""),
        decision.confidence if decision else 0,
        _text(facts.setup_type),
        _text(facts.setup_grade),
        facts.pivot_price,
        facts.entry_price,
        facts.stop_price,
        facts.available_r_multiple,
        facts.relative_strength_percentile,
        facts.sector_name,
        facts.sector_rank or "N/A",
        facts.industry_group or "N/A",
        facts.industry_group_rank or "N/A",
        _text(facts.volume_grade),
        _text(facts.volume_profile.volume_state),
        _text(facts.cpr_profile.cpr_state),
        _text(facts.avwap_profile.state),
        _text(facts.risk_grade),
        _joined(decision.warnings if decision else ()),
    ]


def _top_sheet(workbook: Workbook, results: Sequence[StockResult], generated: datetime) -> None:
    """Create the branded Top 20 dashboard and ranked table."""
    sheet = workbook.create_sheet("Top 20")
    _title(
        sheet,
        "TRADERSHIVE",
        f"EOD MOMENTUM SCANNER  •  Generated {generated:%d-%b-%Y %H:%M %Z}",
        len(TOP_HEADERS),
    )
    situation = results[0].situation_profile
    decision = results[0].decision_profile
    actions = [
        result.decision_profile.action.value if result.decision_profile else ""
        for result in results
    ]
    breadth = (
        situation.breadth_profile.breadth_state.value
        if situation and situation.breadth_profile
        else "Unavailable"
    )
    kpis = [
        ("Scan Date", generated.date()),
        ("Market Regime", _text(situation.market_regime if situation else None)),
        ("Trading Bias", _text(situation.trading_bias if situation else None)),
        ("Aggression", _text(situation.aggression if situation else None)),
        ("Risk Environment", _text(situation.risk_environment if situation else None)),
        ("Position Sizing", _text(situation.position_sizing_guidance if situation else None)),
        ("Breadth State", breadth),
        ("CPR State", situation.cpr_environment if situation else "Unavailable"),
        ("Scanner Profile", decision.scanner_profile if decision else "Momentum Breakout"),
        ("Symbols Scanned", len(results)),
        ("BUY", actions.count("BUY")),
        ("WATCHLIST", actions.count("WATCHLIST")),
        ("AVOID", actions.count("AVOID")),
    ]
    for index, (label, value) in enumerate(kpis):
        _kpi(sheet, 4 + (index // 7) * 3, 1 + (index % 7) * 3, label, value)
        sheet.merge_cells(
            start_row=4 + (index // 7) * 3,
            start_column=1 + (index % 7) * 3,
            end_row=4 + (index // 7) * 3,
            end_column=2 + (index % 7) * 3,
        )
        sheet.merge_cells(
            start_row=5 + (index // 7) * 3,
            start_column=1 + (index % 7) * 3,
            end_row=5 + (index // 7) * 3,
            end_column=2 + (index % 7) * 3,
        )
    sheet["A4"].number_format = "dd-mmm-yyyy"
    start_row = 10
    rows = [_top_row(result) for result in results[:20]]
    last_row = _table(sheet, start_row, TOP_HEADERS, rows, "Top20Results")
    _format_columns(sheet, start_row, last_row, TOP_HEADERS)
    _widths(sheet, TOP_HEADERS, {"Warnings"})
    sheet.column_dimensions["A"].width = 8
    sheet.freeze_panes = f"C{start_row + 1}"
    sheet.auto_filter.ref = f"A{start_row}:W{last_row}"
    sheet.print_area = f"A1:W{last_row}"
    sheet.print_title_rows = f"1:{start_row}"
    _top_conditional_formatting(sheet, start_row + 1, last_row)
    _configure_sheet(sheet, f"1:{start_row}")


def _top_conditional_formatting(sheet: Worksheet, first_row: int, last_row: int) -> None:
    """Apply actionable color cues to the Top 20 table."""
    if last_row < first_row:
        return
    fills = {"BUY": GREEN, "WATCHLIST": PALE_YELLOW, "AVOID": RED}
    for value, color in fills.items():
        sheet.conditional_formatting.add(
            f"C{first_row}:C{last_row}",
            FormulaRule(
                formula=[f'C{first_row}="{value}"'], fill=PatternFill("solid", fgColor=color)
            ),
        )
    grade_colors = {
        "A+": DARK_GREEN,
        "A": GREEN,
        "B": PALE_YELLOW,
        "C": ORANGE,
        "D": RED,
        "Reject": RED,
    }
    for column in ("E", "V"):
        for value, color in grade_colors.items():
            sheet.conditional_formatting.add(
                f"{column}{first_row}:{column}{last_row}",
                FormulaRule(
                    formula=[f'{column}{first_row}="{value}"'],
                    fill=PatternFill("solid", fgColor=color),
                ),
            )
    for column in ("D", "M"):
        sheet.conditional_formatting.add(
            f"{column}{first_row}:{column}{last_row}",
            DataBarRule(
                start_type="num", start_value=0, end_type="num", end_value=100, color=YELLOW
            ),
        )
    available = f"L{first_row}:L{last_row}"
    rules = [
        ("lessThan", 2, RED),
        ("between", [2, 3], ORANGE),
        ("between", [3, 5], PALE_YELLOW),
        ("greaterThanOrEqual", 5, GREEN),
    ]
    for operator, formula, color in rules:
        formulas = formula if isinstance(formula, list) else [formula]
        sheet.conditional_formatting.add(
            available,
            CellIsRule(
                operator=operator, formula=formulas, fill=PatternFill("solid", fgColor=color)
            ),
        )
    sheet.conditional_formatting.add(
        f"Q{first_row}:Q{last_row}",
        FormulaRule(
            formula=[f'OR(Q{first_row}="N/A",Q{first_row}=0)'],
            fill=PatternFill("solid", fgColor=MID_GREY),
            font=Font(color=DARK_GREY),
        ),
    )


def _situation_sheet(
    workbook: Workbook, results: Sequence[StockResult], generated: datetime
) -> None:
    """Create the first-sheet market and leadership dashboard."""
    sheet = workbook.create_sheet("Situation Summary", 0)
    _title(sheet, "TRADERSHIVE", f"SITUATION SUMMARY  •  {generated:%d-%b-%Y %H:%M %Z}", 12)
    situation = results[0].situation_profile
    if situation is None:
        sheet["A4"] = "Situation profile unavailable"
        sheet["A4"].fill = PatternFill("solid", fgColor=PALE_YELLOW)
        _configure_sheet(sheet)
        return
    breadth = situation.breadth_profile
    cards = [
        ("Market Regime", situation.market_regime.value),
        ("Market Health", situation.market_health),
        ("Trading Bias", situation.trading_bias.value),
        ("Aggression", situation.aggression.value),
        ("Breadth", breadth.breadth_state.value if breadth else "Unavailable"),
        ("Money Flow", situation.money_flow.value),
        ("Risk Environment", situation.risk_environment.value),
        ("Position Sizing", situation.position_sizing_guidance.value),
        ("Max Positions", situation.recommended_maximum_open_positions),
        ("Max Risk / Trade", situation.maximum_risk_per_trade / 100),
        ("Holding Period", situation.expected_holding_period),
        ("CPR Environment", situation.cpr_environment),
    ]
    for index, (label, value) in enumerate(cards):
        row = 4 + (index // 4) * 3
        column = 1 + (index % 4) * 3
        _kpi(sheet, row, column, label, value)
        sheet.merge_cells(start_row=row, start_column=column, end_row=row, end_column=column + 1)
        sheet.merge_cells(
            start_row=row + 1, start_column=column, end_row=row + 1, end_column=column + 1
        )
    sheet["J8"].number_format = "0.00%"
    _section(sheet, 14, 1, 6, "Market Environment & Breadth")
    environment = [
        ("Breadth Score", breadth.score if breadth else None),
        ("Breadth Confidence", breadth.confidence if breadth else None),
        ("CPR Breakout Participation", situation.cpr_breakout_participation),
        ("AVWAP Environment", situation.avwap_environment.value),
        ("AVWAP Support Participation", situation.avwap_support_participation),
        ("Market Pressure", situation.market_pressure_score),
        ("Risk-On State", situation.risk_on_state.value),
    ]
    for offset, (label, value) in enumerate(environment, 15):
        sheet.cell(offset, 1, label).font = Font(name=FONT_NAME, bold=True, color=CHARCOAL)
        sheet.cell(offset, 2, value)
    _section(sheet, 14, 7, 12, "Risk & Strategy Guidance")
    guidance = [
        ("Recommended Setups", ", ".join(item.value for item in situation.recommended_setup_types)),
        ("Recommended Strategies", _joined(situation.recommended_strategy)),
        ("Avoid / Warnings", _joined(situation.warnings) or "None"),
        ("Reasons", _joined(situation.reasons)),
    ]
    for offset, (label, value) in enumerate(guidance, 15):
        sheet.cell(offset, 7, label).font = Font(name=FONT_NAME, bold=True, color=CHARCOAL)
        sheet.merge_cells(start_row=offset, start_column=8, end_row=offset, end_column=12)
        sheet.cell(offset, 8, value).alignment = Alignment(wrap_text=True, vertical="top")
        sheet.row_dimensions[offset].height = 32
    _leadership_tables(sheet, results, 24)
    for column in range(1, 13):
        sheet.column_dimensions[get_column_letter(column)].width = 15
    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["G"].width = 23
    sheet.freeze_panes = "A4"
    sheet.print_area = f"A1:L{sheet.max_row}"
    _configure_sheet(sheet, "1:3")


def _section(sheet: Worksheet, row: int, start_column: int, end_column: int, label: str) -> None:
    """Add a dark section banner outside tabular data."""
    sheet.merge_cells(start_row=row, start_column=start_column, end_row=row, end_column=end_column)
    cell = sheet.cell(row, start_column, label)
    cell.fill = PatternFill("solid", fgColor=CHARCOAL)
    cell.font = Font(name=FONT_NAME, bold=True, color=YELLOW)
    cell.alignment = Alignment(vertical="center")


def _leadership_tables(sheet: Worksheet, results: Sequence[StockResult], start_row: int) -> None:
    """Render sector and available industry leadership without synthesizing metrics."""
    situation = results[0].situation_profile
    leadership = situation.sector_leadership if situation else None
    sector_lists = [
        ("Top 5 Sectors", leadership.top_sectors if leadership else ()),
        ("Bottom 5 Sectors", leadership.bottom_sectors if leadership else ()),
        ("Improving Sectors", leadership.improving_sectors if leadership else ()),
        ("Weakening Sectors", leadership.weakening_sectors if leadership else ()),
    ]
    for index, (label, values) in enumerate(sector_lists):
        column = 1 + index * 3
        _section(sheet, start_row, column, column + 1, label)
        for offset, value in enumerate(tuple(values)[:5], start_row + 1):
            sheet.cell(offset, column, offset - start_row)
            sheet.cell(offset, column + 1, value)
    profiles: dict[str, Any] = {}
    for result in results:
        if result.facts.industry_group and result.facts.industry_group_profile:
            profiles[result.facts.industry_group] = result.facts.industry_group_profile
    ranked = sorted(profiles.items(), key=lambda item: item[1].rank)
    industry_lists = [
        ("Top 5 Industry Groups", ranked[:5]),
        ("Bottom 5 Industry Groups", list(reversed(ranked[-5:]))),
        (
            "Improving Groups",
            [item for item in ranked if _text(item[1].rotation) == "Improving"][:5],
        ),
        (
            "Weakening Groups",
            [item for item in ranked if _text(item[1].rotation) == "Weakening"][:5],
        ),
    ]
    industry_row = start_row + 8
    for index, (label, values) in enumerate(industry_lists):
        column = 1 + index * 3
        _section(sheet, industry_row, column, column + 1, label)
        for offset, (name, profile) in enumerate(values, industry_row + 1):
            sheet.cell(offset, column, profile.rank)
            sheet.cell(offset, column + 1, name)


def _detailed_rows(results: Sequence[StockResult]) -> tuple[list[str], list[list[Any]]]:
    """Build the complete candidate audit payload."""
    headers = [
        "Rank",
        "Symbol",
        "Action",
        "Decision Score",
        "Decision Confidence",
        "Decision Grade",
        "Profile",
        "Decision Reasons",
        "Warnings",
        "Market Score",
        "Sector",
        "Sector Score",
        "Sector Rank",
        "Sector Rotation",
        "Industry Group",
        "Industry Score",
        "Industry Rank",
        "Industry Percentile",
        "Industry Rotation",
        "RS Percentile",
        "RS Profile",
        "Stock Score",
        "Stock Grade",
        "Setup Type",
        "Setup Score",
        "Setup Confidence",
        "Setup Grade",
        "Pattern Score",
        "Structure Score",
        "Compression Score",
        "Setup Volume Score",
        "Location Score",
        "Advanced Setup Score",
        "Base Maturity Score",
        "Failure Risk Score",
        "Prior Advance Score",
        "Catalyst Quality Score",
        "Breakout Retest Score",
        "Stage Quality Score",
        "Setup Reasons",
        "Setup Warnings",
        "Volume Score",
        "Volume Confidence",
        "Volume Grade",
        "Volume State",
        "Volume Signature",
        "Accumulation Score",
        "Distribution Score",
        "Dry-Up Score",
        "Expansion Score",
        "Volume Quality Score",
        "Pocket Pivot",
        "Volume Reasons",
        "Volume Warnings",
        "CPR Score",
        "CPR Grade",
        "CPR State",
        "Breakout Probability %",
        "Trend Probability %",
        "Range Probability %",
        "AVWAP Score",
        "AVWAP Grade",
        "AVWAP State",
        "AVWAP Support Score",
        "AVWAP Resistance Score",
        "AVWAP Alignment Score",
        "AVWAP Anchors",
        "Risk Score",
        "Risk Confidence",
        "Risk Grade",
        "Stop Quality Score",
        "Volatility Score",
        "Liquidity Score",
        "Reward Score",
        "Extension Score",
        "Pivot",
        "Entry",
        "Stop",
        "Invalidation",
        "Nearest Resistance",
        "Available R",
        "Stop Distance %",
        "Stop ATR",
        "Target 2R",
        "Target 3R",
        "Target 4R",
        "Target 5R",
        "Risk Reasons",
        "Risk Warnings",
        "Legacy Scanner Score",
        "Breadth Grade",
        "Breadth Score",
        "Breadth State",
        "Breadth Confidence",
        "RS Feature Score",
        "Trend Feature Score",
        "Volume Feature Score",
        "Momentum Feature Score",
        "Volatility Feature Score",
    ]
    rows: list[list[Any]] = []
    for result in results:
        facts = result.facts
        decision = result.decision_profile
        setup = facts.setup_profile
        volume = facts.volume_profile
        cpr = facts.cpr_profile
        avwap = facts.avwap_profile
        risk = facts.risk_profile
        industry = facts.industry_group_profile
        rs = facts.rs_profile
        rs_profile = (
            ", ".join(
                f"{horizon}D: {value.relative_return:.2f}"
                for horizon, value in rs.horizons().items()
            )
            if rs
            else "N/A"
        )
        anchors = (
            ", ".join(f"{key}: {value}" for key, value in avwap.anchor_summary.items()) or "N/A"
        )
        rows.append(
            [
                result.rank,
                result.symbol,
                _text(decision.action if decision else None, ""),
                result.decision_score,
                decision.confidence if decision else 0,
                _text(decision.grade if decision else None, ""),
                decision.scanner_profile if decision else "Momentum Breakout",
                _joined(decision.reasons if decision else ()),
                _joined(decision.warnings if decision else ()),
                result.features["market"].score,
                facts.sector_name,
                result.features["sector"].score,
                facts.sector_rank or "N/A",
                _text(facts.sector_rotation),
                facts.industry_group or "N/A",
                facts.industry_group_score if industry else None,
                facts.industry_group_rank or "N/A",
                facts.industry_group_percentile if industry else None,
                _text(facts.industry_group_rotation) if industry else "N/A",
                facts.relative_strength_percentile,
                rs_profile,
                facts.stock_score,
                _text(facts.stock_grade),
                _text(setup.best_setup_type),
                setup.score,
                setup.confidence,
                _text(setup.grade),
                setup.pattern_score,
                setup.structure_score,
                setup.compression_score,
                setup.volume_score,
                setup.location_score,
                setup.advanced_setup_score,
                setup.base_maturity_score,
                setup.failure_risk_score,
                setup.prior_advance_score,
                setup.catalyst_quality_score,
                setup.breakout_retest_score,
                setup.stage_quality_score,
                _joined(setup.reasons),
                _joined(setup.warnings),
                volume.score,
                volume.confidence,
                _text(volume.grade),
                _text(volume.volume_state),
                _text(volume.volume_signature),
                volume.accumulation_score,
                volume.distribution_score,
                volume.dryup_score,
                volume.expansion_score,
                volume.quality_score,
                volume.pocket_pivot,
                _joined(volume.reasons),
                _joined(volume.warnings),
                cpr.score,
                _text(cpr.grade),
                _text(cpr.cpr_state),
                cpr.breakout_probability,
                cpr.trend_probability,
                cpr.range_probability,
                avwap.score,
                _text(avwap.grade),
                _text(avwap.state),
                avwap.support_score,
                avwap.resistance_score,
                avwap.alignment_score,
                anchors,
                risk.score,
                risk.confidence,
                _text(risk.grade),
                risk.stop_quality_score,
                risk.volatility_score,
                risk.liquidity_score,
                risk.reward_score,
                risk.extension_score,
                facts.pivot_price,
                facts.entry_price,
                facts.stop_price,
                facts.invalidation_price,
                risk.facts.nearest_resistance,
                facts.available_r_multiple,
                risk.facts.stop_distance_percent,
                risk.facts.stop_distance_atr,
                risk.facts.target_2r,
                risk.facts.target_3r,
                risk.facts.target_4r,
                risk.facts.target_5r,
                _joined(risk.reasons),
                _joined(risk.warnings),
                result.final_score,
                _text(facts.breadth_grade),
                facts.breadth_score,
                _text(facts.breadth_profile.breadth_state),
                facts.breadth_profile.confidence,
                result.features["relative_strength"].score,
                result.features["trend"].score,
                result.features["volume"].score,
                result.features["momentum"].score,
                result.features["volatility"].score,
            ]
        )
    return headers, rows


def _data_sheet(
    workbook: Workbook,
    title: str,
    subtitle: str,
    headers: list[str],
    rows: list[list[Any]],
    table_name: str,
    generated: datetime,
    long_columns: set[str] | None = None,
) -> Worksheet:
    """Create one consistently themed intelligence table sheet."""
    sheet = workbook.create_sheet(title)
    _title(
        sheet,
        title.upper(),
        f"{subtitle}  •  Generated {generated:%d-%b-%Y %H:%M %Z}",
        len(headers),
    )
    last_row = _table(sheet, 4, headers, rows, table_name)
    _format_columns(sheet, 4, last_row, headers)
    _widths(sheet, headers, long_columns)
    for column, header in enumerate(headers, 1):
        letter = get_column_letter(column)
        data_range = f"{letter}5:{letter}{last_row}"
        sheet.conditional_formatting.add(
            data_range,
            FormulaRule(
                formula=[f'{letter}5="N/A"'],
                fill=PatternFill("solid", fgColor=MID_GREY),
                font=Font(color=DARK_GREY),
            ),
        )
        if "Warning" in header:
            sheet.conditional_formatting.add(
                data_range,
                FormulaRule(
                    formula=[f'{letter}5<>""'],
                    fill=PatternFill("solid", fgColor=PALE_YELLOW),
                ),
            )
    sheet.print_area = f"A1:{get_column_letter(len(headers))}{last_row}"
    _configure_sheet(sheet, "1:4")
    return sheet


def _specialist_sheets(
    workbook: Workbook, results: Sequence[StockResult], generated: datetime
) -> None:
    """Create focused views over the unchanged scanner result payload."""
    detailed_headers, detailed_rows = _detailed_rows(results)
    detailed = workbook.create_sheet("Detailed Candidates")
    _title(
        detailed,
        "DETAILED CANDIDATES",
        f"FULL CANDIDATE EVIDENCE  •  Generated {generated:%d-%b-%Y %H:%M %Z}",
        len(detailed_headers),
    )
    groups = [
        ("Decision & Context", 1, 9),
        ("Leadership", 10, 19),
        ("Stock & Setup", 20, 41),
        ("Volume", 42, 54),
        ("CPR & AVWAP", 55, 67),
        ("Risk & Trade Plan", 68, len(detailed_headers) - 10),
        ("Compatibility Fields", len(detailed_headers) - 9, len(detailed_headers)),
    ]
    for label, start, end in groups:
        _section(detailed, 4, start, end, label)
    detailed_last_row = _table(detailed, 5, detailed_headers, detailed_rows, "DetailedCandidates")
    _format_columns(detailed, 5, detailed_last_row, detailed_headers)
    for column, header in enumerate(detailed_headers, 1):
        letter = get_column_letter(column)
        data_range = f"{letter}6:{letter}{detailed_last_row}"
        detailed.conditional_formatting.add(
            data_range,
            FormulaRule(
                formula=[f'{letter}6="N/A"'],
                fill=PatternFill("solid", fgColor=MID_GREY),
                font=Font(color=DARK_GREY),
            ),
        )
        if "Warning" in header:
            detailed.conditional_formatting.add(
                data_range,
                FormulaRule(
                    formula=[f'{letter}6<>""'],
                    fill=PatternFill("solid", fgColor=PALE_YELLOW),
                ),
            )
    _widths(
        detailed,
        detailed_headers,
        {
            "Decision Reasons",
            "Warnings",
            "Setup Reasons",
            "Setup Warnings",
            "Volume Reasons",
            "Volume Warnings",
            "AVWAP Anchors",
            "Risk Reasons",
            "Risk Warnings",
            "RS Profile",
        },
    )
    detailed.freeze_panes = "C6"
    detailed.print_title_rows = "1:5"
    detailed.print_area = f"A1:{get_column_letter(len(detailed_headers))}{detailed_last_row}"
    _configure_sheet(detailed, "1:5")

    specs = [
        (
            "Advanced Setup Details",
            [
                "Rank",
                "Symbol",
                "Setup Type",
                "Setup Grade",
                "Setup Score",
                "Setup Confidence",
                "Pattern Score",
                "Structure Score",
                "Compression Score",
                "Setup Volume Score",
                "Location Score",
                "Advanced Setup Score",
                "Base Maturity Score",
                "Failure Risk Score",
                "Prior Advance Score",
                "Catalyst Quality Score",
                "Breakout Retest Score",
                "Stage Quality Score",
                "Pivot",
                "Invalidation",
                "Setup Reasons",
                "Setup Warnings",
            ],
        ),
        (
            "Volume Intelligence",
            [
                "Rank",
                "Symbol",
                "Volume Score",
                "Volume Confidence",
                "Volume Grade",
                "Volume State",
                "Volume Signature",
                "Accumulation Score",
                "Distribution Score",
                "Dry-Up Score",
                "Expansion Score",
                "Volume Quality Score",
                "Pocket Pivot",
                "Volume Reasons",
                "Volume Warnings",
            ],
        ),
        (
            "Sector and Industry Leadership",
            [
                "Rank",
                "Symbol",
                "Sector",
                "Sector Score",
                "Sector Rank",
                "Sector Rotation",
                "Industry Group",
                "Industry Score",
                "Industry Rank",
                "Industry Percentile",
                "Industry Rotation",
                "RS Percentile",
            ],
        ),
        (
            "CPR",
            [
                "Rank",
                "Symbol",
                "CPR Score",
                "CPR Grade",
                "CPR State",
                "Breakout Probability %",
                "Trend Probability %",
                "Range Probability %",
            ],
        ),
        (
            "AVWAP",
            [
                "Rank",
                "Symbol",
                "AVWAP Score",
                "AVWAP Grade",
                "AVWAP State",
                "AVWAP Support Score",
                "AVWAP Resistance Score",
                "AVWAP Alignment Score",
                "AVWAP Anchors",
            ],
        ),
        (
            "Risk",
            [
                "Rank",
                "Symbol",
                "Risk Score",
                "Risk Confidence",
                "Risk Grade",
                "Stop Quality Score",
                "Volatility Score",
                "Liquidity Score",
                "Reward Score",
                "Extension Score",
                "Entry",
                "Stop",
                "Nearest Resistance",
                "Available R",
                "Stop Distance %",
                "Stop ATR",
                "Target 2R",
                "Target 3R",
                "Target 4R",
                "Target 5R",
                "Risk Reasons",
                "Risk Warnings",
            ],
        ),
    ]
    indices = {header: index for index, header in enumerate(detailed_headers)}
    for number, (title, headers) in enumerate(specs, 1):
        rows = [[row[indices[header]] for header in headers] for row in detailed_rows]
        _data_sheet(
            workbook,
            title,
            "SCANNER INTELLIGENCE",
            headers,
            rows,
            f"Intel{number}",
            generated,
            {
                name
                for name in headers
                if "Reason" in name or "Warning" in name or "Anchors" in name
            },
        )
    validation_headers = ["Metric", "Value", "Notes"]
    actions = [
        result.decision_profile.action.value if result.decision_profile else ""
        for result in results
    ]
    validation_rows = [
        [
            "Report Type",
            "Live EOD scan",
            "Historical validation metrics are not part of live scanner results",
        ],
        [
            "Scanner Profile",
            (
                results[0].decision_profile.scanner_profile
                if results[0].decision_profile
                else "Momentum Breakout"
            ),
            "Applied by the scanner",
        ],
        ["Symbols Scanned", len(results), "Complete eligible result set"],
        ["BUY Candidates", actions.count("BUY"), "Current live classifications"],
        ["WATCHLIST Candidates", actions.count("WATCHLIST"), "Current live classifications"],
        ["AVOID Candidates", actions.count("AVOID"), "Current live classifications"],
    ]
    _data_sheet(
        workbook,
        "Validation Summary",
        "LIVE REPORT METADATA",
        validation_headers,
        validation_rows,
        "ValidationMetadata",
        generated,
        {"Notes"},
    )


def generate_excel_report(results: Sequence[StockResult], destination: Path) -> Path:
    """Write ranked results to an atomic, professionally formatted workbook."""
    if not results:
        raise ValueError("Cannot generate a report without results")
    destination = destination.resolve()
    destination.parent.mkdir(parents=True, exist_ok=True)
    temporary = destination.with_suffix(f"{destination.suffix}.tmp")
    generated = datetime.now().astimezone()

    workbook = Workbook()
    workbook.remove(workbook.active)
    _situation_sheet(workbook, results, generated)
    _top_sheet(workbook, results, generated)
    _specialist_sheets(workbook, results, generated)
    workbook.properties.title = "TradersHIVE EOD Momentum Scanner"
    workbook.properties.subject = "Presentation-ready EOD scan intelligence"
    workbook.properties.creator = "TradersHIVE"
    workbook.properties.created = generated.replace(tzinfo=None)
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.save(temporary)
    temporary.replace(destination)
    return destination
