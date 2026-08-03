"""JSON, CSV, and multi-sheet Excel historical validation reports."""

import csv
from collections.abc import Mapping, Sequence
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from app.validation.models import HistoricalTrade, ValidationMetrics, ValidationReport


def write_json_report(report: ValidationReport, destination: Path) -> Path:
    """Serialize a validation report atomically as readable JSON."""
    destination.parent.mkdir(parents=True, exist_ok=True)
    temporary = destination.with_suffix(destination.suffix + ".tmp")
    temporary.write_text(report.model_dump_json(indent=2), encoding="utf-8")
    temporary.replace(destination)
    return destination


def write_trade_ledger(trades: Sequence[HistoricalTrade], destination: Path) -> Path:
    """Write the complete candidate and outcome ledger as CSV."""
    destination.parent.mkdir(parents=True, exist_ok=True)
    rows = [_trade_row(trade) for trade in trades]
    headers = list(rows[0]) if rows else ["scan_date", "symbol", "outcome"]
    with destination.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)
    return destination


def write_excel_report(report: ValidationReport, destination: Path) -> Path:
    """Create the ten-sheet validation workbook required for review."""
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    _metrics_sheet(workbook, "Summary", {"Overall": report.metrics})
    mappings = (
        ("Score Buckets", "score_buckets"),
        ("Setup Types", "setup_types"),
        ("Market Regimes", "market_regimes"),
        ("Sectors", "sectors"),
        ("Annual Results", "calendar_years"),
    )
    for title, key in mappings:
        _metrics_sheet(workbook, title, report.breakdowns[key].groups)
    _metrics_sheet(workbook, "Top-N Comparison", report.top_n_comparison)
    for title, key in (
        ("Advanced Setups", "advanced_setups"),
        ("Volume Signatures", "volume_signatures"),
        ("Market Pressure", "market_pressure"),
        ("Industry Groups", "industry_groups"),
        ("Scanner Profiles", "scanner_profiles"),
    ):
        _metrics_sheet(workbook, title, report.breakdowns[key].groups)
    feature = workbook.create_sheet("Feature Analysis")
    feature.append(["Feature", "Sample Size", "Trigger Rate", "Win Rate", "Expectancy", "95% CI"])
    for name, analysis in report.feature_analysis.items():
        feature.append(
            [
                name,
                analysis.sample_size,
                analysis.trigger_rate,
                analysis.metrics.win_rate,
                analysis.metrics.expectancy,
                str(analysis.win_rate_confidence_interval or "Insufficient sample"),
            ]
        )
    ledger = workbook.create_sheet("Trade Ledger")
    rows = [_trade_row(trade) for trade in report.trades]
    if rows:
        ledger.append(list(rows[0]))
        for row in rows:
            ledger.append(list(row.values()))
    else:
        ledger.append(["No trades"])
    walk = workbook.create_sheet("Walk Forward")
    walk.append(["In Sample", "Out of Sample", "IS Expectancy", "OOS Expectancy"])
    if report.walk_forward:
        for period in report.walk_forward.periods:
            walk.append(
                [
                    f"{period.in_sample_start} to {period.in_sample_end}",
                    f"{period.out_of_sample_start} to {period.out_of_sample_end}",
                    period.in_sample_metrics.expectancy,
                    period.out_of_sample_metrics.expectancy,
                ]
            )
    quality = workbook.create_sheet("Data Quality Warnings")
    quality.append(["Data Quality Score", report.data_quality.score])
    quality.append(["Warning"])
    for warning in report.data_quality.warnings:
        quality.append([warning])
    for sheet in workbook.worksheets:
        for cell in sheet[1]:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = PatternFill("solid", fgColor="17365D")
        sheet.freeze_panes = "A2"
    temporary = destination.with_suffix(destination.suffix + ".tmp")
    workbook.save(temporary)
    temporary.replace(destination)
    return destination


def _metrics_sheet(workbook: Workbook, title: str, groups: Mapping[str, ValidationMetrics]) -> None:
    sheet = workbook.create_sheet(title)
    fields = list(ValidationMetrics.model_fields)
    sheet.append(["Group", *fields])
    for name, metrics in groups.items():
        values = metrics.model_dump()
        sheet.append([name, *(values[field] for field in fields)])


def _trade_row(trade: HistoricalTrade) -> dict[str, object]:
    candidate = trade.candidate.model_dump(mode="json")
    outcome = trade.model_dump(mode="json", exclude={"candidate"})
    return {**candidate, **outcome}
